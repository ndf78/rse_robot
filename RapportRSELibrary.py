import openpyxl
import os

def creer_le_fichier_de_rapport(fichier_sortie):
    classeur = openpyxl.Workbook()
    classeur.save(fichier_sortie)
    classeur.close()
    return "Le fichier de sortie a été crée a l'emplacement" + fichier_sortie

def initialiser_tableau_cas_de_test(fichier_sortie, titre_cas_de_test):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    i = 1
    start = feuille.cell(1, 1).value
    if start == None:
        feuille.cell(1, 1).value = titre_cas_de_test
        feuille.cell(2, 1).value = "URL"
        feuille.cell(2, 2).value = "Propreté (en %)"
        feuille.cell(2, 3).value = "Consommation de CO2 (en g)"
    else:
        while start != None:
            start1 = feuille.cell(i, 1).value
            start2 = feuille.cell(i + 1, 1).value
            if start1 == start2:
                start = None
            else:
                start = "value"
            i += 1
        feuille.cell(i, 1).value = titre_cas_de_test
        feuille.cell(1 + i, 1).value = "URL"
        feuille.cell(1 + i, 2).value = "Propreté (en %)"
        feuille.cell(1 + i, 3).value = "Consommation de CO2 (en g)"
    classeur.save(fichier_sortie)
    classeur.close()
    return

def renseigner_tableau_cas_de_test(fichier_sortie, url, proprete, consommation_co2):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    i = 1
    start = feuille.cell(1, 1).value
    while start != None:
        start1 = feuille.cell(i, 1).value
        start2 = feuille.cell(i + 1, 1).value
        if start1 == start2:
            start = None
        else:
            start = "value"
        i += 1
    feuille.cell(i - 1, 1).value = url
    feuille.cell(i - 1, 2).value = proprete
    feuille.cell(i - 1, 3).value = consommation_co2
    classeur.save(fichier_sortie)
    classeur.close()
    return

def initialiser_tableau_cas_de_test_total(fichier_sortie):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    feuille.cell(1, 5).value = "Cas de tests"
    feuille.cell(1, 6).value = "Consommation C02 total du cas de test (en g)"
    classeur.save(fichier_sortie)
    classeur.close()    
    return

def renseigner_tableau_cas_de_test_total(fichier_sortie, titre_cas_de_test, consommation_co2_total):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    i = 1
    start = feuille.cell(1, 1).value
    while start != None:
        start1 = feuille.cell(i, 5).value
        start2 = feuille.cell(i + 1, 5).value
        if start1 == start2:
            start = None
        else:
            start = "value"
        i += 1
    feuille.cell(i - 1, 5).value = titre_cas_de_test
    feuille.cell(i - 1, 6).value = consommation_co2_total
    classeur.save(fichier_sortie)
    classeur.close()
    return

def initialiser_tableau_campagne_de_tests(fichier_sortie):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    feuille.cell(1, 8).value = "Campagne de tests"
    feuille.cell(1, 9).value = "Consommation CO2 total de la campagne de test (en g)"
    classeur.save(fichier_sortie)
    classeur.close()
    return

def renseigner_tableau_campagne_de_tests(fichier_sortie, nom_de_la_campagne_de_tests, consommation_de_la_campagne_de_tests):
    classeur = openpyxl.load_workbook(fichier_sortie)
    feuille = classeur.active
    feuille.cell(2, 8).value = nom_de_la_campagne_de_tests
    feuille.cell(2, 9).value = consommation_de_la_campagne_de_tests
    classeur.save(fichier_sortie)
    classeur.close()
    return

def executer_lighthouse(url, sortie):
    cmd = os.system('lighthouse {} --quiet --chrome-flags="--headless" --disable-storage-reset --output json --output html --output-path {}'.format(url, sortie))
    return cmd

def executer_lighthouse_avec_identifiant(url, option, sortie):
    cmd = os.system('lighthouse {} {} --output json --output html --output-path {}'.format(url, option, sortie))
    return cmd

# url = "https://www.ausy.fr/fr/nos-solutions/"
# url2 = "www.ausy.fr_fr_nos-solutions_"
# sortie = "/home/ndf/github/rse_robot/rapport/{}".format(url2)
# executer_lighthouse(url, sortie)


# creer_le_fichier_de_rapport("rapport/rapport.xlsx")
# initialiser_tableau_cas_de_test("rapport/rapport.xlsx", "Cas de test 56")
# renseigner_tableau_cas_de_test('rapport/rapport.xlsx',"lapin","catapulte","pierre")
# initialiser_tableau_cas_de_test_total('rapport/rapport.xlsx')
# renseigner_tableau_cas_de_test_total('rapport/rapport.xlsx', "Cas de test 18", 0.58)
# initialiser_tableau_campagne_de_tests('rapport/rapport.xlsx')
# renseigner_tableau_campagne_de_tests("rapport/rapport.xlsx", "RSE", 58.2)